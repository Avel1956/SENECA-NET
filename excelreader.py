import streamlit as st

import pandas as pd
from pandas_profiling import ProfileReport
import seaborn as sns
import matplotlib.pyplot as plt
import os
import openpyxl
from streamlit_pandas_profiling import st_profile_report

def read_excel_file(file_path):
    wb_obj = openpyxl.load_workbook(file_path)
    sheets = wb_obj.sheetnames
    sheet = st.selectbox("Seleccione hoja", sheets)
    sheet_obj = wb_obj[sheet]
    # ask for the range of cells to read
    start_row = st.number_input("Fila inicial", min_value=1, value=1)
    start_col = st.number_input("Columna inicial", min_value=1, value=1)
    end_row = st.number_input("Fila final", min_value=1, value=1)
    end_col = st.number_input("Columna final", min_value=1, value=1)
    # read the data
    data = sheet_obj.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col, values_only=True)

   
    
    cols = next(data)[0:]
    df = pd.DataFrame(data, columns=cols)
    return df

def show_basic_statistics(df):
    st.write("### Estadísticas básicas")
    st.write(df.describe())

def select_analysis_type():
    st.write("### Analysis Type")
    analysis_type = st.selectbox(
        "Select analysis type",
        ["Una columna", "Dos columnas", "Varias columnas"]
    )
    return analysis_type

def select_single_column(df):
    st.write("### Analisis de una columna")
    column_name = st.selectbox("Seleccione columna", df.columns)
    st.write("### Analisis disponibles")
    analysis = st.selectbox(
        "Seleccione analisis",
        ["Histograma", "Densidad", "Box Plot"]
    )
    return column_name, analysis

def select_two_columns(df):
    st.write("### Analisis de dos columnas")
    x_col = st.selectbox("Selecione columna X", df.columns)
    y_col = st.selectbox("Selecione columna Y", df.columns)
    st.write("### Analisis disponbiles")
    analysis = st.selectbox(
        "Seleccione analisis",
        ["Dispersión", "Linea"]
    )
    return x_col, y_col, analysis

def select_several_columns(df):
    st.write("### Analisis de varias columnas")
    columns = st.multiselect("Seleccione columnas", df.columns)
    st.write("### Analisis disponibles")
    analysis = st.selectbox(
        "Seleccione analisis",
        ["Plot por pares", "Heatmap"]
    )
    return columns, analysis

def generate_chart(df, analysis_type, options):
    st.write("### Figura")
    if analysis_type == "Una columna":
        column_name, analysis = options
        if analysis == "Histograma":
            fig, ax = plt.subplots()
            sns.histplot(df[column_name], ax=ax)
            st.pyplot(fig)
        elif analysis == "Densitad":
            fig, ax = plt.subplots()
            sns.kdeplot(df[column_name], ax=ax)
            st.pyplot(fig)
        elif analysis == "Box Plot":
            fig, ax = plt.subplots()
            sns.boxplot(x=df[column_name], ax=ax)
            st.pyplot(fig)
    elif analysis_type == "Dos columnas":
        x_col, y_col, analysis = options
        if analysis == "Dispersión":
            fig, ax = plt.subplots()
            sns.scatterplot(x=x_col, y=y_col, data=df, ax=ax)
            st.pyplot(fig)
        elif analysis == "Linea":
            fig, ax = plt.subplots()
            sns.lineplot(x=x_col, y=y_col, data=df, ax=ax)
            st.pyplot(fig)
    elif analysis_type == "Varias columnas":
        columns, analysis = options
        if analysis == "Plot por pares":
            fig = sns.pairplot(df[columns])
            st.pyplot(fig)
        elif analysis == "Heatmap":
            fig, ax = plt.subplots()
            sns.heatmap(df[columns].corr(), annot=True, ax=ax)
            st.pyplot(fig)

def main():
    st.title("Analisis de archivos Excel")
    file_path = st.file_uploader("Suba el archivo", type=["xlsx"])
    if file_path is not None:
        df = read_excel_file(file_path)
        show_basic_statistics(df)
        st.write("### Perfilamiento")
        profile = st.checkbox("Hacer perfilamiento")
        if profile:
            pr = ProfileReport(df)
            st_profile_report(pr)
        analysis_type = select_analysis_type()
        if analysis_type == "Una columna":
            options = select_single_column(df)
        elif analysis_type == "Dos columnas":
            options = select_two_columns(df)
        elif analysis_type == "Varias columnas":
            options = select_several_columns(df)
        generate_chart(df, analysis_type, options)
        st.write("### Guardar figura")
        chart_name = st.text_input("Entre el nombre de la figura")
        if chart_name:
            chart_path = f"{chart_name}.html"
            st.write(f"Guardando figura en {chart_path}")
            plt.savefig(chart_path, bbox_inches="tight")
        st.write("### Reiniciar la aplicación")
        restart = st.button("Presione para reiniciar")
        if restart:
            os._exit(0)

if __name__ == "__main__":
    main()
