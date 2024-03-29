import streamlit as st
import datetime
import pandas as pd
from pandas_profiling import ProfileReport
from streamlit_pandas_profiling import st_profile_report
import seaborn as sns
import matplotlib.pyplot as plt
import os
import openpyxl


class ReportSection:
    def __init__(self, name, options, data):
        self.name = name
        #self.description = description
        self.options = options
        self.result = None
        self.figure = None
        self.data = data
        self.created_at = datetime.datetime.now()

def create_report_section(name,  operation, data):
    section = ReportSection(name, operation, data)
    return section

def read_excel_file(file_path):
    wb_obj = openpyxl.load_workbook(file_path)
    sheets = wb_obj.sheetnames
    sheet = st.selectbox("Seleccione hoja", sheets)
    sheet_obj = wb_obj[sheet]
    # ask for the range of cells to read
    start_row = st.number_input("Fila inicial", min_value=1, value=1)
    start_col = st.number_input("Columna inicial", min_value=1, value=1)
    end_row = st.number_input("Fila final", min_value=1, value=1)
    end_col = st.number_input("Columna final", min_value=1, value=1)
    # read the data
    data = sheet_obj.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col, values_only=True)

   
    
    cols = next(data)[0:]
    df = pd.DataFrame(data, columns=cols)
    st.write("### Estado del dataset")
    dft_copy = df.copy()
    # all the columns are string
    for col in dft_copy.columns:
        dft_copy[col] = dft_copy[col].astype(str)
    # check if there are empty cells
    
    st.write(dft_copy.head())
    st.write("### Valores nulos")
    st.write(df.isnull().sum())
    st.write("### Valores duplicados")
    st.write(df.duplicated().sum())
    st.write("### Valores unicos")
    st.write(df.nunique())
    
    return df
# def to clean the data
def clean_data(df):
    # Select box to select the columns to clean
    st.write("### Seleccione columnas a limpiar")
    columns_to_clean = st.multiselect("Seleccione columnas", df.columns)
    # Select box to select the cleaning method
    st.write("### Seleccione metodo de limpieza")
    cleaning_method = st.selectbox(
        "Seleccione metodo de limpieza",
        ["Eliminar filas vacias", "Sustituir valores vacios con NAN","Eliminar columnas", "Eliminar filas", "Eliminar duplicados", "Unify types"]
    
    )
    # Apply the cleaning method
    if cleaning_method == "Eliminar filas vacias":
        df.dropna(inplace=True)
    elif cleaning_method == "Sustituir valores vacios con NAN":
        df.fillna("NAN", inplace=True)
    elif cleaning_method == "Eliminar columnas":
        df.drop(columns_to_clean, axis=1, inplace=True)
    elif cleaning_method == "Eliminar filas":
        df.drop(columns_to_clean, axis=0, inplace=True)
    elif cleaning_method == "Eliminar duplicados":
        df.drop_duplicates(inplace=True)
    elif cleaning_method == "Unify types":
        df[columns_to_clean] = df[columns_to_clean].astype(str)
    st.write("### Estado del dataset")
    st.write(df.head())
    return df
    

    
def show_basic_statistics(df):
    st.write("### Estadísticas básicas")
    st.write(df.describe())

def select_analysis_type():
    st.write("### Analysis Type")
    analysis_type = st.selectbox(
        "Select analysis type",
        ["Una columna", "Dos columnas", "Varias columnas"]
    )
    return analysis_type

def select_single_column(df):
    st.write("### Analisis de una columna")
    column_name = st.selectbox("Seleccione columna", df.columns)
    st.write("### Analisis disponibles")
    analysis = st.selectbox(
        "Seleccione analisis",
        ["Histograma", "Densidad", "Box Plot"]
    )
    return column_name, analysis

def select_two_columns(df):
    st.write("### Analisis de dos columnas")
    x_col = st.selectbox("Selecione columna X", df.columns)
    y_col = st.selectbox("Selecione columna Y", df.columns)
    st.write("### Analisis disponbiles")
    analysis = st.selectbox(
        "Seleccione analisis",
        ["Dispersión", "Linea"]
    )
    return x_col, y_col, analysis

def select_several_columns(df):
    st.write("### Analisis de varias columnas")
    columns = st.multiselect("Seleccione columnas", df.columns)
    st.write("### Analisis disponibles")
    analysis = st.selectbox(
        "Seleccione analisis",
        ["Plot por pares", "Heatmap"]
    )
    return columns, analysis

def generate_chart(df, analysis_type, options):
    st.write("### Figura")
    if analysis_type == "Una columna":
        column_name, analysis = options
        if analysis == "Histograma":
            fig, ax = plt.subplots()
            sns.histplot(df[column_name], ax=ax)
            st.pyplot(fig)
        elif analysis == "Densitad":
            fig, ax = plt.subplots()
            sns.kdeplot(df[column_name], ax=ax)
            st.pyplot(fig)
        elif analysis == "Box Plot":
            fig, ax = plt.subplots()
            sns.boxplot(x=df[column_name], ax=ax)
            st.pyplot(fig)
    elif analysis_type == "Dos columnas":
        x_col, y_col, analysis = options
        if analysis == "Dispersión":
            fig, ax = plt.subplots()
            sns.scatterplot(x=x_col, y=y_col, data=df, ax=ax)
            st.pyplot(fig)
        elif analysis == "Linea":
            fig, ax = plt.subplots()
            sns.lineplot(x=x_col, y=y_col, data=df, ax=ax)
            st.pyplot(fig)
    elif analysis_type == "Varias columnas":
        columns, analysis = options
        if analysis == "Plot por pares":
            fig = sns.pairplot(df[columns])
            st.pyplot(fig)
        elif analysis == "Heatmap":
            fig, ax = plt.subplots()
            sns.heatmap(df[columns].corr(), annot=True, ax=ax)
            st.pyplot(fig)



def main():

    
    st.title("Herramienta de analisis de datos")
    file_path = st.file_uploader("Suba el archivo", type=["xlsx"])
    if file_path is not None:
        df = read_excel_file(file_path)
        show_basic_statistics(df)
        st.write("### Perfilamiento")
        profile = st.checkbox("Hacer perfilamiento")
        if profile:
            pr = ProfileReport(df, title="Reporte de dataset")
            st_profile_report(pr)
            #dialogo para descargar el reporte como html
            if st.button("Descargar reporte"):
                pr.to_file("reporte.html")
                with open("reporte.html", "rb") as f:
                    file_bytes = f.read()
                st.download_button(
                    label="Descargar reporte",
                    data=file_bytes,
                    file_name="reporte.html",
                    mime="text/html"
                )
                

        
            
        analysis_type = select_analysis_type()
        if analysis_type == "Una columna":
            options = select_single_column(df)
        elif analysis_type == "Dos columnas":
            options = select_two_columns(df)
        elif analysis_type == "Varias columnas":
            options = select_several_columns(df)
        generate_chart(df, analysis_type, options)
        st.write("### Guardar figura")
        chart_name = st.text_input("Entre el nombre de la figura")
        if chart_name:
            chart_path = f"{chart_name}.png"
            st.write(f"Guardando figura en {chart_path}")
            plt.savefig(chart_path, bbox_inches="tight")
        
        st.write("### Guardar analisis")
       
    

        st.write("### Reiniciar la aplicación")
        restart = st.button("Presione para reiniciar")
        if restart:
            os._exit(0)
    with st.sidebar:
        st.header("Saved Sections")
        saved_sections = []  # moved here to keep track of saved sections
        sections_container = st.empty()  # container to show the saved sections
        
        if st.button("Save Analysis"):
            section_name = analysis_type
                
            section_op = options
            section_data = df # or whichever data you want to save
                
            new_section = create_report_section(section_name, section_op, section_data)
            saved_sections.append(new_section)

        # Show the saved sections
        if saved_sections:
            st.subheader("Reorder, Edit or Delete sections:")
            for i, section in enumerate(saved_sections):
                # Create a card for each section
                card = st.container()
                with card:
                    st.write(f"{i+1}. {section.name}")
                    st.write(f"Description: {section.options}")
                    st.write(f"Created At: {section.created_at}")
                    if st.button("Edit"):
                        # Show inputs to edit the section and update it in the list
                        section_name = st.text_input("Enter section name:", section.name)
                        section_op = st.text_input("Enter analysis operation:", section.operation)
                        section.data = st.text_input("Enter data:", section.data)
                        section.created_at = datetime.datetime.now()
                        saved_sections[i] = ReportSection(section_name, section_op, section.data)
                    if st.button("Delete"):
                        # Remove the section from the list of saved sections
                        saved_sections.pop(i)
                        break

                # Reorder the sections
                new_order = st.sidebar.selectbox("Reorder sections:", [i+1 for i in range(len(saved_sections))], index=None)
                if new_order:
                    saved_sections = [saved_sections[i-1] for i in new_order]
                    
                # Update the container with the saved sections
                sections_container.write("Saved Sections:")
                for i, section in enumerate(saved_sections):
                    sections_container.write(f"{i+1}. {section.name}")

        else:
            st.subheader("No saved sections yet.")  # if there are no saved sections
if __name__ == "__main__":
    main()

